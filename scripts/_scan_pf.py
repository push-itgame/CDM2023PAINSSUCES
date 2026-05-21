"""Scan Phase Finale + Grille for input cells."""
import json
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

p = Path(r"c:\Users\mar_q\OneDrive\Documents\GitHub\CDM2023PAINSSUCES\CDM2026_Inscription.xlsx")
wb = openpyxl.load_workbook(p, data_only=True)

for sheet in ["Phase Finale", "Grille"]:
    ws = wb[sheet]
    print("\n===", sheet, "non-empty sample ===")
    n = 0
    for ri in range(1, 220):
        for ci in range(1, 25):
            v = ws.cell(ri, ci).value
            if v is None or str(v).strip() == "":
                continue
            s = str(v).strip()
            if sheet == "Phase Finale" and ri > 50:
                print(f"  {get_column_letter(ci)}{ri}={s[:50]!r}")
                n += 1
            elif sheet == "Grille" and (ri >= 150 or "M7" in s or "match" in s.lower()):
                print(f"  {get_column_letter(ci)}{ri}={s[:50]!r}")
                n += 1
            if n > 60:
                break
        if n > 60:
            break

# Phase finale: find cells with data validation / inputs - scan row 4-50 cols C-P
ws = wb["Phase Finale"]
print("\n=== Phase Finale grid area C4:P54 ===")
for ri in range(4, 55):
    parts = []
    for ci in range(3, 17):
        v = ws.cell(ri, ci).value
        if v is not None and str(v).strip():
            parts.append(f"{get_column_letter(ci)}{ri}={str(v)[:25]!r}")
    if parts:
        print("R"+str(ri), " | ".join(parts[:8]))

wb.close()
