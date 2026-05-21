import openpyxl
from openpyxl.utils import get_column_letter

p = r"c:\Users\mar_q\OneDrive\Documents\GitHub\CDM2023PAINSSUCES\CDM2026_Inscription.xlsx"

# Phase Finale bracket cells (data_only)
wb = openpyxl.load_workbook(p, data_only=True)
pf = wb["Phase Finale"]
print("=== Phase Finale bracket picks (cols C,F,I,L,O,R) ===")
bracket_rows = list(range(5, 68, 2))  # 5,7,9,...,67
slots = [("C", "r32_left"), ("F", "r16"), ("I", "qf"), ("L", "sf"), ("O", "final"), ("R", "winner")]
for ri in bracket_rows:
    vals = []
    for col, _ in slots:
        ci = ord(col) - ord("A") + 1
        v = pf.cell(ri, ci).value
        if v:
            vals.append(f"{col}{ri}={v!r}")
    if vals:
        print(f"R{ri}: " + " | ".join(vals))

# Also check adjacent rows for teams (6,8,10...)
print("\n=== Phase Finale even rows (team B side?) ===")
for ri in range(6, 68, 2):
    vals = []
    for col in ["C", "F", "I", "L", "O"]:
        ci = ord(col) - ord("A") + 1
        v = pf.cell(ri, ci).value
        if v:
            vals.append(f"{col}{ri}={v!r}")
    if vals:
        print(f"R{ri}: " + " | ".join(vals))

# Bonus T column
print("\n=== Bonus T57-T68 ===")
for ri in range(57, 69):
    t = pf.cell(ri, 20).value
    p = pf.cell(ri, 16).value
    if t or p:
        print(f"R{ri} P={p!r} T={t!r}")

# Grille KO scores rows 78-109
wg = wb["Grille"]
print("\n=== Grille KO rows 78-109 all cols ===")
for ri in range(78, 110):
    parts = []
    for ci in range(1, 13):
        v = wg.cell(ri, ci).value
        if v is not None and str(v).strip() != "":
            parts.append(f"{get_column_letter(ci)}{ri}={v!r}")
    if parts:
        print(" | ".join(parts))

# Grille M6:N77 lookup table for bonus
print("\n=== Grille M6:N20 (bonus lookup) ===")
for ri in range(6, 25):
    print(f"R{ri} M={wg.cell(ri,13).value!r} N={wg.cell(ri,14).value!r}")

wb.close()

# Formulas for Grille H78-I109
wb2 = openpyxl.load_workbook(p, data_only=False)
wg2 = wb2["Grille"]
print("\n=== Grille KO formulas H/I cols ===")
for ri in range(78, 110):
    h = wg2.cell(ri, 8).value
    i = wg2.cell(ri, 9).value
    c = wg2.cell(ri, 3).value
    d = wg2.cell(ri, 4).value
    if h or i or c or d:
        print(f"R{ri} C={str(c)[:25]!r} D={d!r} H={str(h)[:40] if h else None!r} I={str(i)[:40] if i else None!r}")
wb2.close()

# Poules - find where user picks winners for KO (X column?)
wb3 = openpyxl.load_workbook(p, data_only=False)
ws = wb3["Poules"]
print("\n=== Poules cols around V-Z row 162 (3rd place) ===")
for ri in range(160, 165):
    parts = []
    for ci in range(20, 31):
        v = ws.cell(ri, ci).value
        if v:
            parts.append(f"{get_column_letter(ci)}{ri}={str(v)[:40]!r}")
    if parts:
        print(" | ".join(parts))
wb3.close()
