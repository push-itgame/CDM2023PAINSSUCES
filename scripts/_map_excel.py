"""Cartographie automatique du template Excel."""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "CDM2026_Inscription.xlsx"
OUT = Path(__file__).resolve().parent / "excel_cell_map.json"

GROUPS = {
    "A": [[1, "Mexique", "Afrique du Sud"], [2, "Corée du Sud", "Rép. Tchèque"]],
}


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Poules"]
    matches = []
    for ri in range(1, 250):
        b = ws.cell(ri, 2).value
        if b is None:
            continue
        try:
            mid = int(b)
        except (TypeError, ValueError):
            continue
        if not (1 <= mid <= 72):
            continue
        matches.append({
            "match": mid,
            "row": ri,
            "home": ws.cell(ri, 9).value,  # I
            "away": ws.cell(ri, 12).value,  # L
            "scoreHome": ws.cell(ri, 10).value,  # J
            "scoreAway": ws.cell(ri, 11).value,  # K
        })
    ident = {}
    for addr, key in [("J2", "nom_complet"), ("J3", "email"), ("J4", "equipe")]:
        ident[key] = {"cell": addr, "value": ws[addr].value}
    wb.close()

    wb2 = openpyxl.load_workbook(XLSX, data_only=False)
    ws_f = wb2["Poules"]
    sample_formula_j = ws_f.cell(10, 10).value
    wb2.close()

    # Phase Finale + Grille quick scan
    wb3 = openpyxl.load_workbook(XLSX, data_only=True)
    pf = wb3["Phase Finale"]
    grille = wb3["Grille"]
    pf_bonus = []
    for ri in range(50, 80):
        label = pf.cell(ri, 15).value  # O
        if label and isinstance(label, str) and len(label) > 5:
            pf_bonus.append({"row": ri, "label": label.strip(), "value_col_p": pf.cell(ri, 16).value})
    grille_inputs = []
    for ri in range(150, 190):
        label = grille.cell(ri, 3).value
        if label and isinstance(label, str):
            grille_inputs.append({
                "row": ri,
                "label": label.strip(),
                "value_d": grille.cell(ri, 4).value,
            })
    wb3.close()

    data = {
        "template": str(XLSX.name),
        "sheets": ["Reglement", "Poules", "Phase Finale", "Grille"],
        "identite": ident,
        "poules_score_cols": {"scoreHome": "J", "scoreAway": "K", "homeTeam": "I", "awayTeam": "L", "matchNum": "B"},
        "poules_matches": matches,
        "sample_formula_note": str(sample_formula_j)[:120] if sample_formula_j else "",
        "phase_finale_bonus_hints": pf_bonus,
        "grille_hints": grille_inputs,
    }
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print("Wrote", OUT)
    print("Matches mapped:", len(matches))
    print("Ident:", ident)


if __name__ == "__main__":
    main()
