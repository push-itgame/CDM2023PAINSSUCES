#!/usr/bin/env python3
"""Convertit CDM2026_Inscription.xlsx en JSON grille (format site CDM 2026)."""
from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Dépendance manquante : pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = Path(__file__).resolve().parent
CELL_MAP = SCRIPTS / "excel_cell_map.json"
BRACKET_MAP = SCRIPTS / "excel_bracket_map.json"
RESULTATS = ROOT / "data" / "resultats.json"

SKIP_VALUES = frozenset({"", "#N/A", "#REF!", "#VALUE!", None})


def load_json(path: Path) -> dict:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def strip_accents(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )


def build_team_aliases() -> dict[str, str]:
    """Mappe variantes Excel → noms officiels (depuis resultats.json)."""
    aliases: dict[str, str] = {}
    if RESULTATS.exists():
        data = load_json(RESULTATS)
        for m in (data.get("matchs") or {}).values():
            for key in ("home", "away"):
                name = (m.get(key) or "").strip()
                if name:
                    aliases[_alias_key(name)] = name
    # Variantes fréquentes dans le template Excel
    manual = {
        "etats unis": "États-Unis",
        "etats-unis": "États-Unis",
        "haiti": "Haïti",
        "bosnie herzegovine": "Bosnie Herzégovine",
        "coree du sud": "Corée du Sud",
        "rep. tcheque": "Rép. Tchèque",
        "rep tcheque": "Rép. Tchèque",
        "cote d'ivoire": "Côte d'Ivoire",
        "cote divoire": "Côte d'Ivoire",
        "ecosse": "Écosse",
        "egypte": "Égypte",
        "equateur": "Équateur",
        "mexique": "Mexique",
        "bresil": "Brésil",
        "senegal": "Sénégal",
        "nouvelle zelande": "Nouvelle Zélande",
        "arabie saoudite": "Arabie Saoudite",
        "ouzbekistan": "Ouzbékistan",
        "curacao": "Curacao",
        "rd congo": "RD Congo",
        "turquie": "Turquie",
        "panama": "Panama",
        "paraguay": "Paraguay",
        "colombie": "Colombie",
        "autriche": "Autriche",
        "australie": "Australie",
        "japon": "Japon",
        "maroc": "Maroc",
        "qatar": "Qatar",
        "suisse": "Suisse",
        "canada": "Canada",
        "france": "France",
        "allemagne": "Allemagne",
        "espagne": "Espagne",
        "portugal": "Portugal",
        "angleterre": "Angleterre",
        "belgique": "Belgique",
        "pays-bas": "Pays-Bas",
        "pays bas": "Pays-Bas",
        "croatie": "Croatie",
        "suede": "Suède",
        "tunisie": "Tunisie",
        "algerie": "Algérie",
        "argentine": "Argentine",
        "uruguay": "Uruguay",
        "iran": "Iran",
        "irak": "Irak",
        "norvege": "Norvège",
        "ghana": "Ghana",
        "cap vert": "Cap Vert",
        "afrique du sud": "Afrique du Sud",
    }
    for k, v in manual.items():
        aliases.setdefault(k, v)
    return aliases


def _alias_key(name: str) -> str:
    s = strip_accents(name.strip().lower())
    s = re.sub(r"[^\w\s'-]", "", s)
    return re.sub(r"\s+", " ", s).strip()


TEAM_ALIASES = build_team_aliases()


def normalize_team(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s or s in SKIP_VALUES:
        return ""
    key = _alias_key(s)
    return TEAM_ALIASES.get(key, s)


def normalize_bonus(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s or s in SKIP_VALUES:
        return ""
    if s.lower() == "autre":
        return "Autre"
    return s


def cell_str(ws, addr: str):
    if not addr:
        return None
    return ws[addr].value


def _ident_cell(ident_cfg: dict, key: str, default: str) -> str:
    raw = ident_cfg.get(key, default)
    if isinstance(raw, dict):
        return raw.get("cell", default)
    return raw or default


def read_identite(ws_poules, ident_cfg: dict) -> dict:
    nom_complet = (cell_str(ws_poules, _ident_cell(ident_cfg, "nom_complet", "J2")) or "").strip()
    parts = nom_complet.split(None, 1) if nom_complet else []
    prenom = parts[0] if parts else ""
    nom = parts[1] if len(parts) > 1 else ""
    return {
        "prenom": prenom,
        "nom": nom,
        "email": (cell_str(ws_poules, _ident_cell(ident_cfg, "email", "J3")) or "").strip(),
        "equipe": (cell_str(ws_poules, _ident_cell(ident_cfg, "equipe", "J4")) or "").strip(),
    }


def score_value(raw) -> str | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, float) and raw.is_integer():
        return str(int(raw))
    return str(raw).strip()


def read_poules_matches(ws_poules, poules_matches: list) -> dict:
    matchs = {}
    for entry in poules_matches:
        mid = entry["match"]
        home = normalize_team(entry.get("home") or ws_poules.cell(entry["row"], 9).value)
        away = normalize_team(entry.get("away") or ws_poules.cell(entry["row"], 12).value)
        sh = score_value(ws_poules.cell(entry["row"], 10).value)
        sa = score_value(ws_poules.cell(entry["row"], 11).value)
        if sh is None and sa is None:
            continue
        matchs[f"Match {mid}"] = {
            "home": home,
            "away": away,
            "scoreHome": sh if sh is not None else "",
            "scoreAway": sa if sa is not None else "",
        }
    return matchs


def read_col_values(ws, col: str, rows: list[int], *, teams: bool = True) -> list[str]:
    out = []
    norm = normalize_team if teams else normalize_bonus
    for row in rows:
        v = ws[f"{col}{row}"].value
        s = norm(v)
        if s:
            out.append(s)
    return out


def read_bracket(ws_pf, bracket_cfg: dict) -> tuple[dict, dict, dict, dict]:
    pf = bracket_cfg["phase_finale"]

    r32 = read_col_values(ws_pf, pf["r32_teams_col"], pf["r32_teams_rows"])
    r16 = read_col_values(ws_pf, pf["r32_winners_col"], pf["r32_winners_rows"])
    qf = read_col_values(ws_pf, pf["r16_winners_col"], pf["r16_winners_rows"])
    sf = read_col_values(ws_pf, pf["qf_winners_col"], pf["qf_winners_rows"])
    finalists = read_col_values(ws_pf, pf["sf_winners_col"], pf["sf_winners_rows"])
    winner = normalize_team(ws_pf[f"{pf['winner_col']}{pf['winner_row']}"].value)

    etape2 = {
        "r32": r32,
        "r16": r16,
        "qf": qf,
        "sf": sf,
        "finalists": finalists,
        "winner": winner,
    }

    bareme = {
        "liste32QualifiesIssuesPoules": r32,
        "vainqueursSeiziemePourHuitiemes16": r16,
        "vainqueursHuitiemesPourQuarts8": qf,
        "vainqueursQuartsPourDemis4": sf,
        "finalistesChoisis": finalists,
        "troisiemePlaceChoix": "",
        "vainqueurFinal": winner,
    }

    etape2_pick: dict[str, str] = {}
    vainqueurs: dict[str, str] = {}

    for mid, row in zip(pf["r32_match_order"], pf["r32_winners_rows"], strict=False):
        val = normalize_team(ws_pf[f"{pf['r32_winners_col']}{row}"].value)
        if val:
            etape2_pick[str(mid)] = val
            vainqueurs[str(mid)] = val

    for mid, row in zip(pf["r16_match_order"], pf["r16_winners_rows"], strict=False):
        val = normalize_team(ws_pf[f"{pf['r16_winners_col']}{row}"].value)
        if val:
            etape2_pick[str(mid)] = val

    for mid, row in zip(pf["qf_match_order"], pf["qf_winners_rows"], strict=False):
        val = normalize_team(ws_pf[f"{pf['qf_winners_col']}{row}"].value)
        if val:
            etape2_pick[str(mid)] = val

    for mid, row in zip(pf["sf_match_order"], pf["sf_winners_rows"], strict=False):
        val = normalize_team(ws_pf[f"{pf['sf_winners_col']}{row}"].value)
        if val:
            etape2_pick[str(mid)] = val

    if winner:
        etape2_pick[str(pf["final_match"])] = winner

    return etape2, bareme, etape2_pick, vainqueurs


def read_bonus(ws_pf, ws_grille, bonus_cfg: dict) -> dict:
    bonus = {}
    sheet_name = bonus_cfg.get("sheet", "Phase Finale")
    ws = ws_pf if sheet_name == "Phase Finale" else ws_grille
    col_i = bonus_cfg.get("grille_fallback_col", "I")
    fallback_rows = bonus_cfg.get("grille_fallback_rows") or {}

    for key, addr in (bonus_cfg.get("cells") or {}).items():
        val = normalize_bonus(cell_str(ws, addr))
        if not val and key in fallback_rows:
            row = fallback_rows[key]
            val = normalize_bonus(ws_grille[f"{col_i}{row}"].value)
        bonus[key] = val
    return bonus


def convert(xlsx_path: Path, *, cell_map_path: Path = CELL_MAP, bracket_map_path: Path = BRACKET_MAP) -> dict:
    cell_map = load_json(cell_map_path)
    bracket_map = load_json(bracket_map_path)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws_poules = wb["Poules"]
    ws_pf = wb["Phase Finale"]
    ws_grille = wb["Grille"]

    identite = read_identite(ws_poules, bracket_map.get("identite") or cell_map.get("identite", {}))
    matchs = read_poules_matches(ws_poules, cell_map.get("poules_matches") or [])
    etape2, bareme, etape2_pick, vainqueurs = read_bracket(ws_pf, bracket_map)
    bonus = read_bonus(ws_pf, ws_grille, bracket_map.get("bonus") or {})
    wb.close()

    payload = {
        "identite": identite,
        "matchs": matchs,
        "equipesQualifiees32Liste": bareme["liste32QualifiesIssuesPoules"],
        "etape2Tableau": etape2,
        "etape2Pick": etape2_pick,
        "vainqueursTableauEliminationChoisis": vainqueurs,
        "phaseFinalePourBareme": bareme,
        "bonus": bonus,
        "scoresElimination": {},
        "sourceExcel": xlsx_path.name,
    }
    return payload


def default_out_path(payload: dict, out_dir: Path) -> Path:
    ident = payload.get("identite") or {}
    prenom = (ident.get("prenom") or "Joueur").strip().replace(" ", "_")
    nom = (ident.get("nom") or "Anonyme").strip().replace(" ", "_")
    return out_dir / f"CDM2026_{prenom}_{nom}.json"


def main() -> int:
    parser = argparse.ArgumentParser(description="Excel CDM 2026 → JSON grille")
    parser.add_argument("xlsx", type=Path, help="Fichier CDM2026_Inscription.xlsx rempli")
    parser.add_argument("--out", type=Path, help="Fichier JSON de sortie")
    parser.add_argument("--out-dir", type=Path, default=ROOT / "data" / "grilles")
    parser.add_argument("--dry-run", action="store_true", help="Affiche un résumé sans écrire")
    args = parser.parse_args()

    if not args.xlsx.is_file():
        print(f"Fichier introuvable : {args.xlsx}", file=sys.stderr)
        return 1

    try:
        payload = convert(args.xlsx)
    except KeyError as e:
        print(f"Feuille Excel manquante : {e}", file=sys.stderr)
        return 1

    ident = payload["identite"]
    if not ident.get("email"):
        print("Avertissement : e-mail vide (obligatoire pour Supabase).", file=sys.stderr)

    summary = {
        "joueur": f"{ident.get('prenom', '')} {ident.get('nom', '')}".strip(),
        "email": ident.get("email", ""),
        "matchs_poules": len(payload["matchs"]),
        "equipes_32": len(payload["equipesQualifiees32Liste"]),
        "picks_tableau": len(payload["etape2Pick"]),
        "bonus_remplis": sum(1 for v in payload["bonus"].values() if v),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))

    if args.dry_run:
        return 0

    out_path = args.out or default_out_path(payload, args.out_dir)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
        f.write("\n")
    print(f"Écrit : {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
