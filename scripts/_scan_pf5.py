import openpyxl
from openpyxl.utils import get_column_letter

p = r"c:\Users\mar_q\OneDrive\Documents\GitHub\CDM2023PAINSSUCES\CDM2026_Inscription.xlsx"
wb = openpyxl.load_workbook(p, data_only=False)
pf = wb["Phase Finale"]

# Full row dump for bracket area
for ri in [5, 6, 7, 8, 9, 10, 11, 12, 45, 46, 47, 48, 57, 58, 59, 60, 63, 64, 65, 66, 67, 68]:
    parts = []
    for ci in range(1, 19):
        v = pf.cell(ri, ci).value
        if v is not None and str(v).strip():
            parts.append(f"{get_column_letter(ci)}{ri}={str(v)[:60]!r}")
    print(f"R{ri}: " + " | ".join(parts))

# Grille row 110-172 H column refs
wg = wb["Grille"]
print("\n=== Grille bracket summary H col 110-172 ===")
for ri in range(110, 173):
    h = wg.cell(ri, 8).value
    c = wg.cell(ri, 3).value
    if h or c:
        print(f"R{ri} C={str(c)[:35] if c else ''!r} H={str(h)[:50] if h else ''!r}")

# Poules X column - what is user input?
ws = wb["Poules"]
print("\n=== Poules row 11 formulas X,Y,Z ===")
for ri in [11, 12, 22, 23, 55, 56, 162, 163]:
    parts = []
    for ci in range(21, 31):
        v = ws.cell(ri, ci).value
        if v:
            parts.append(f"{get_column_letter(ci)}{ri}={str(v)[:50]!r}")
    print(" | ".join(parts))

# Find data validation / input cells - search for cells that reference T57 bonus input
print("\n=== Phase Finale T col formulas ===")
for ri in range(55, 69):
    t = pf.cell(ri, 20).value
    p = pf.cell(ri, 16).value
    s = pf.cell(ri, 19).value
    if t or p or s:
        print(f"R{ri} P={p!r} S={s!r} T={t!r}")

# Grille bonus I173-I182 formulas  
print("\n=== Grille bonus I173-I182 ===")
for ri in range(173, 183):
    i = wg.cell(ri, 9).value
    h = wg.cell(ri, 8).value
    print(f"R{ri} H={str(h)[:40] if h else None!r} I={str(i)[:50] if i else None!r}")

wb.close()
