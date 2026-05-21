import openpyxl
from openpyxl.utils import get_column_letter

p = r"c:\Users\mar_q\OneDrive\Documents\GitHub\CDM2023PAINSSUCES\CDM2026_Inscription.xlsx"
wb = openpyxl.load_workbook(p, data_only=False)
ws = wb["Grille"]
print("=== Grille rows 1-30 ===")
for ri in range(1, 31):
    parts = []
    for ci in range(1, 12):
        v = ws.cell(ri, ci).value
        if v is not None and str(v).strip():
            parts.append(f"{get_column_letter(ci)}{ri}={str(v)[:35]!r}")
    if parts:
        print("R"+str(ri), " | ".join(parts))

print("\n=== Grille rows 130-190 (values col D,E,H,I) ===")
for ri in range(130, 195):
    c = ws.cell(ri, 3).value
    d = ws.cell(ri, 4).value
    h = ws.cell(ri, 8).value
    i = ws.cell(ri, 9).value
    if c or d or (h and h != 0) or i:
        print(f"R{ri} C={str(c)[:40] if c else ''!r} D={d!r} H={h!r} I={str(i)[:30] if i else ''!r}")

# Find KO match score rows - search M73 pattern
print("\n=== Search M7 / match KO ===")
for ri in range(1, 150):
    for ci in range(1, 15):
        v = ws.cell(ri, ci).value
        if v and isinstance(v, str) and ("M7" in v or "M10" in v or "Seizi" in v):
            print(get_column_letter(ci)+str(ri), repr(v)[:60])
wb.close()

wb2 = openpyxl.load_workbook(p, data_only=True)
ws2 = wb2["Grille"]
print("\n=== Grille data D173-D182 (bonus?) ===")
for ri in range(170, 185):
    print(f"R{ri} C={ws2.cell(ri,3).value!r} D={ws2.cell(ri,4).value!r}")
wb2.close()
