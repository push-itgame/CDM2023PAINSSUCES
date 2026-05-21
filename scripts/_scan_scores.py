import openpyxl
from openpyxl.utils import get_column_letter

p = r"c:\Users\mar_q\OneDrive\Documents\GitHub\CDM2023PAINSSUCES\CDM2026_Inscription.xlsx"
wb = openpyxl.load_workbook(p, data_only=False)
pf = wb["Phase Finale"]

print("=== Phase Finale F,I,L,O,R sample formulas ===")
for ri in [6, 10, 14, 18, 8, 16, 24, 12, 28, 20, 36, 52]:
    for col in ["D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R"]:
        v = pf[f"{col}{ri}"].value
        if v:
            print(f"{col}{ri}={str(v)[:70]!r}")

# Map C rows to match numbers via Grille
wg = wb["Grille"]
print("\n=== Grille rows 78-109 formulas C,D,H,I ===")
for ri in range(78, 110):
    row = []
    for col in ["C", "D", "E", "F", "G", "H", "I", "J", "K"]:
        v = wg[f"{col}{ri}"].value
        if v:
            row.append(f"{col}{ri}={str(v)[:45]!r}")
    if row:
        print(f"R{ri}: " + " | ".join(row))

# Poules - search for KO score area (rows 175+)
ws = wb["Poules"]
print("\n=== Poules rows 175-220 col B match nums ===")
for ri in range(175, 230):
    b = ws.cell(ri, 2).value
    if b:
        j, k = ws.cell(ri, 10).value, ws.cell(ri, 11).value
        print(f"R{ri} B={b!r} J={j!r} K={k!r} I={ws.cell(ri,9).value!r} L={ws.cell(ri,12).value!r}")

# Search all sheets for literal 73 in col B
for sn in wb.sheetnames:
    ws2 = wb[sn]
    for ri in range(1, 250):
        for ci in [2, 3, 4]:
            v = ws2.cell(ri, ci).value
            if v == 73 or v == "73":
                print(f"Found 73 at {sn}!{get_column_letter(ci)}{ri}")

wb.close()
