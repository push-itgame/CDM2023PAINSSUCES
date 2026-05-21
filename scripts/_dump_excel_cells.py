"""Dump cellules avec coordonnées — python scripts/_dump_excel_cells.py [sheet]"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DEFAULT = ROOT / "CDM2026_Inscription.xlsx"


def dump_sheet(ws, max_row=250, max_col=80):
    for ri in range(1, max_row + 1):
        parts = []
        for ci in range(1, max_col + 1):
            v = ws.cell(ri, ci).value
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            parts.append(f"{get_column_letter(ci)}{ri}={s[:40]!r}")
        if parts:
            print(f"R{ri}:", " | ".join(parts[:15]))
            if len(parts) > 15:
                print("     ...", len(parts) - 15, "more")


def main():
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT
    sheet = sys.argv[2] if len(sys.argv) > 2 else "Poules"
    wb = openpyxl.load_workbook(path, data_only=True)
    print("Sheet:", sheet)
    dump_sheet(wb[sheet])
    wb.close()


if __name__ == "__main__":
    main()
