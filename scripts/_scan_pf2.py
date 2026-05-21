import openpyxl
from openpyxl.utils import get_column_letter

p = r"c:\Users\mar_q\OneDrive\Documents\GitHub\CDM2023PAINSSUCES\CDM2026_Inscription.xlsx"
wb = openpyxl.load_workbook(p, data_only=False)
ws = wb["Phase Finale"]
for ri in range(1, 25):
    parts = []
    for ci in range(1, 20):
        v = ws.cell(ri, ci).value
        if v is not None and str(v).strip():
            s = str(v).replace("\n", " ")[:40]
            parts.append(f"{get_column_letter(ci)}{ri}={s!r}")
    if parts:
        print("R"+str(ri), " | ".join(parts))
print("--- bonus rows ---")
for ri in range(55, 70):
    o = ws.cell(ri, 15).value
    pval = ws.cell(ri, 16).value
    qval = ws.cell(ri, 17).value
    if o:
        print(f"O{ri}={str(o)[:50]!r} P={pval!r} Q={qval!r}")
wb.close()

wb = openpyxl.load_workbook(p, data_only=True)
ws = wb["Phase Finale"]
for ri in range(55, 70):
    o = ws.cell(ri, 15).value
    pval = ws.cell(ri, 16).value
    if o:
        print(f"[data] O{ri} P={pval!r}")
# scan for nation names in bracket area
nations = ["Mexique", "France", "Brésil", "Argentine"]
for ri in range(4, 55):
    for ci in range(3, 16):
        v = ws.cell(ri, ci).value
        if v and isinstance(v, str) and len(v) > 2:
            print(f"PF {get_column_letter(ci)}{ri}={v!r}")
wb.close()
