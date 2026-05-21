"""Exploration one-shot du template Excel — usage: python scripts/_peek_excel.py [path.xlsx]"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT = ROOT / "CDM2026_Inscription.xlsx"


def main() -> int:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT
    if not path.is_file():
        print("Fichier introuvable:", path)
        return 1
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print("Fichier:", path)
    print("Feuilles:", wb.sheetnames)
    keywords = re.compile(
        r"match|groupe|pr[eé]nom|nom|mail|e-?mail|bonus|buteur|joueur|gardien|"
        r"seizi|huit|quart|demi|finale|vainqueur|m7[0-9]|m10[0-4]|m[0-9]{1,2}\b",
        re.I,
    )
    for name in wb.sheetnames:
        ws = wb[name]
        print("\n===", name, "===")
        hits = 0
        for ri, row in enumerate(ws.iter_rows(max_row=400, values_only=True), start=1):
            for ci, cell in enumerate(row, start=1):
                if cell is None:
                    continue
                s = str(cell).strip()
                if not s or not keywords.search(s):
                    continue
                col = openpyxl.utils.get_column_letter(ci)
                print(f"  {col}{ri}: {s[:80]}")
                hits += 1
                if hits > 80:
                    print("  ... (troncature)")
                    break
            if hits > 80:
                break
    wb.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
