import openpyxl
from openpyxl.utils import get_column_letter

p = r"c:\Users\mar_q\OneDrive\Documents\GitHub\CDM2023PAINSSUCES\CDM2026_Inscription.xlsx"
wb = openpyxl.load_workbook(p, data_only=False)
ws = wb["Phase Finale"]
print("=== Phase Finale full scan cols A-T rows 1-70 ===")
for ri in range(1, 71):
    parts = []
    for ci in range(1, 21):
        v = ws.cell(ri, ci).value
        if v is not None and str(v).strip():
            s = str(v)
            if s.startswith("="):
                s = s[:50]
            parts.append(f"{get_column_letter(ci)}{ri}={s!r}")
    if parts:
        print("R"+str(ri), " | ".join(parts[:8]))
        if len(parts) > 8:
            print("   ...", " | ".join(parts[8:]))

# T column bonus inputs
print("\n=== T57-T68 (bonus values) ===")
for ri in range(55, 70):
    print(f"R{ri} T={ws.cell(ri,20).value!r} S={ws.cell(ri,19).value!r} P={ws.cell(ri,16).value!r}")

# Search for score inputs (M73 etc)
print("\n=== Search score / M7 ===")
for ri in range(1, 71):
    for ci in range(1, 21):
        v = ws.cell(ri, ci).value
        if v and (isinstance(v, (int, float)) or (isinstance(v, str) and any(x in str(v) for x in ["M7", "M8", "M9", "score", "Score"]))):
            if not str(v).startswith("=IF"):
                print(get_column_letter(ci)+str(ri), repr(v)[:80])
wb.close()
