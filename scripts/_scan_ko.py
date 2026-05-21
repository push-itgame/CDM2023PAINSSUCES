import openpyxl
from openpyxl.utils import get_column_letter

p = r"c:\Users\mar_q\OneDrive\Documents\GitHub\CDM2023PAINSSUCES\CDM2026_Inscription.xlsx"
wb = openpyxl.load_workbook(p, data_only=True)
ws = wb["Poules"]
print("=== Poules matches 73-104 ===")
for ri in range(1, 250):
    b = ws.cell(ri, 2).value
    if b is None:
        continue
    try:
        mid = int(b)
    except:
        continue
    if 73 <= mid <= 104:
        print(f"M{mid} R{ri} I={ws.cell(ri,9).value!r} L={ws.cell(ri,12).value!r} J={ws.cell(ri,10).value!r} K={ws.cell(ri,11).value!r}")

# Scan V columns for qualified teams (R32)
print("\n=== Poules V/AB/AC/AD cols rows 10-170 (qualifiers) ===")
for ri in [11,12,22,23,33,34,44,45,55,56,66,67,77,78,88,89,99,100,110,111,121,122,132,133,144,155,162]:
    v = ws.cell(ri, 22).value  # V
    if v:
        print(f"R{ri} V={v!r} W={ws.cell(ri,23).value!r} X={ws.cell(ri,24).value!r}")

# Grille rows 78-130 (KO section?)
print("\n=== Grille rows 78-130 ===")
wg = wb["Grille"]
for ri in range(78, 132):
    parts = []
    for ci in range(1, 10):
        v = wg.cell(ri, ci).value
        if v is not None and str(v).strip():
            parts.append(f"{get_column_letter(ci)}{ri}={str(v)[:30]!r}")
    if parts:
        print(" | ".join(parts))
wb.close()
