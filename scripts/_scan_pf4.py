import openpyxl
from openpyxl.utils import get_column_letter

p = r"c:\Users\mar_q\OneDrive\Documents\GitHub\CDM2023PAINSSUCES\CDM2026_Inscription.xlsx"
wb = openpyxl.load_workbook(p, data_only=False)
pf = wb["Phase Finale"]
print("=== Phase Finale R4-R68 (formulas, cols B-S) ===")
for ri in range(4, 69):
    parts = []
    for ci in range(2, 20):
        v = pf.cell(ri, ci).value
        if v is not None and str(v).strip():
            s = str(v).replace("\n", " ")[:55]
            parts.append(f"{get_column_letter(ci)}{ri}={s!r}")
    if parts:
        print(f"R{ri}:")
        for p2 in parts:
            print("  ", p2)

ws = wb["Poules"]
print("\n=== Poules qualification cols U-AD row 10-175 (sample) ===")
for ri in range(10, 176):
    u = ws.cell(ri, 21).value  # U
    v = ws.cell(ri, 22).value  # V - 1st place team name?
    w = ws.cell(ri, 23).value  # W - points?
    x = ws.cell(ri, 24).value  # X - user pick marker?
    y = ws.cell(ri, 25).value
    z = ws.cell(ri, 26).value
    if v and isinstance(v, str) and len(v) > 2 and not str(v).startswith("="):
        print(f"R{ri} U={u!r} V={v!r} W={w!r} X={x!r} Y={y!r} Z={z!r}")

# Search Poules for match 73 or S1
print("\n=== Search 'S1' or 73 in Poules ===")
for ri in range(1, 200):
    for ci in range(1, 30):
        v = ws.cell(ri, ci).value
        if v in (73, "73", "S1", "M73") or (isinstance(v, str) and "Match 73" in v):
            print(f"{get_column_letter(ci)}{ri}={v!r}")

wb.close()
